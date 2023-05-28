import openpyxl, os
os.chdir(r'C:\Users\mohamed alghaly\Desktop')
workbook_to_copy_from = openpyxl.load_workbook('Cast.xlsx')
print("OPENING THE CAST FILE .........")
sheet_from = workbook_to_copy_from.active
data = {}
for i in range(1, 211):
    name = sheet_from.cell(i, 1).value
    actors = sheet_from.cell(i, 2).value
    data[name] = actors
    print(f"the {name}  ------------  {actors} copied")
workbook_to_copy_from.close()
print("THE CAST FILE CLOSED .......   ")

workbook_to_copy_to = openpyxl.load_workbook('Movies.xlsx')
sheet_to = workbook_to_copy_to.active

print("ADDING THE CAST TO THE MOVIES FILE ..........")
for j in range(2, 216):
    name = sheet_to.cell(j, 1).value
    if name in data:
        actors = data[name]
        sheet_to.cell(j, 9).value = actors
        print(f"the {name}  ------------  {actors} pasted")
workbook_to_copy_to.close()
workbook_to_copy_to.save('Movies.xlsx')
print("MOVIES FILE CLOSED ..........")






