import openpyxl


class Movie:
    def __init__(self, name):
        name = str(name)
        self.name = name.title()
        self.director = None
        self.assistant_director = None
        self.time = None
        self.actors = []
        self.rate = None
        self.IMDb_rate = None
        self.year = None
        self.genres = []
        self.votes = None

    def set_time(self, time):
        self.time = time

    def get_time(self):
        return self.time

    def set_directors(self, directors):
        directors = directors.split(',')
        directors = list(map(lambda x: x.title(), directors))
        self.director = directors[0].strip().title()
        if len(directors) - 1:
            self.assistant_director = directors[1].strip().title()

    def get_directors(self):
        return [self.director] if self.assistant_director is None else [self.director, self.assistant_director]

    def set_rate(self, rate):
        self.rate = rate

    def set_IMDb_rate(self, rate):
        self.IMDb_rate = rate

    def get_rate(self):
        return {"me": self.rate, "IMDb": self.IMDb_rate}

    def set_votes(self, votes):
        self.votes = votes

    def get_votes(self):
        return self.votes

    def set_year(self, year):
        self.year = year

    def get_year(self):
        return self.year

    def set_actors(self, actors):
        actors = actors.split(',')
        actors = list(map(lambda x: x.title().strip(), actors))
        self.actors = actors

    def get_actors(self):
        return self.actors

    def set_genres(self, genres):
        genres = genres.split(',')
        genres = list(map(lambda x: x.title().strip(), genres))
        self.genres = genres

    def get_genres(self):
        return self.genres


class Series(Movie):
    episodes = {"Prison Break": 90, 'Friends': 240, 'Breaking Bad': 62, 'Better Call Saul': 50, 'Kidding': 10}

    def get_time(self):
        return self.episodes[self.name] * self.time

    def set_directors(self, directors):
        return

    def get_directors(self):
        return []


class Actor:
    type_ = "Actors"

    def __init__(self, name):
        self.name = name
        self.movies = []

    def set_movies(self, movies):
        movies.sort(key=lambda x: x.get_rate()['me'], reverse=True)
        self.movies = movies

    def get_movies(self):
        movies = map(lambda x: x.name, self.movies)
        return " - ".join(movies)

    def __len__(self):
        return len(self.movies)

    def get_rate(self):
        my_total = sum(i.get_rate()['me'] for i in self.movies)
        IMDb_total = sum(i.get_rate()['IMDb'] for i in self.movies)
        my_total /= len(self)
        my_total = round(my_total, 2)
        IMDb_total /= len(self)
        IMDb_total = round(IMDb_total, 2)
        return {'me': my_total, 'IMDb': IMDb_total}

    def best(self, n=3):
        return self.movies[:n]

    def worst(self, n=3):
        return self.movies[-n:]

    def get_best(self):
        best = self.best()
        best = map(lambda x: x.name, best)
        return " - ".join(best)

    def get_worst(self):
        worst = self.worst()
        worst = map(lambda x: x.name, worst)
        return " - ".join(worst)


class Director(Actor):
    type_ = "Directors"


class Genre(Actor):
    type_ = "Genres"


class Year(Actor):
    type_ = "Years"


watched_movies = []
workbook = openpyxl.load_workbook('Movies.xlsx')
print("OPENING THE MOVIES EXCEL FILE .........")
sheet = workbook.active
NUMBER_OF_MOVIES = sheet.max_row

result = open("Result.txt", 'w')
result2 = open('Result2.txt', 'w')
actors_test = open("Actors.txt", 'w')
directors_test = open('Directors.txt', 'w')
genres_test = open('Genres.txt', 'w')
years_test = open('Years.txt', 'w')

for row in range(2, NUMBER_OF_MOVIES):

    name = sheet.cell(row, 1).value

    if name is None:
        break
        
    print(F"GETTING DATA FOR ----{name}---- MOVIE .......")

    title = sheet.cell(row, 2).value
    IMDb_rate = sheet.cell(row, 3).value
    time = sheet.cell(row, 4).value
    year = sheet.cell(row, 5).value
    genres = sheet.cell(row, 6).value
    votes = sheet.cell(row, 7).value
    actors = sheet.cell(row, 9).value
    if actors is None:
        actors = "Unknown"
    directors = sheet.cell(row, 10).value
    rate = sheet.cell(row, 11).value

    movie = Movie(name) if title == "movie" else Series(name)

    movie.set_actors(actors)
    movie.set_directors(directors)
    movie.set_genres(genres)
    movie.set_rate(rate)
    movie.set_IMDb_rate(IMDb_rate)
    movie.set_time(time)
    movie.set_year(year)
    movie.set_votes(votes)

    print(F"DONE GETTING DATA FOR ----{name}---- MOVIE .......")
    watched_movies.append(movie)

actors = {}
directors = {}
genres = {}
years = {}


def add(dicl, lst, movie):
    for i in lst:
        if i in dicl:
            dicl[i].append(movie)
        else:
            dicl[i] = [movie]
            
            
def test(branch, file):
    for i in branch:
        movies = map(lambda x: x.name, branch[i])
        file.write(f"{i} --> {' - '.join(movies)}\n")


for movie in watched_movies:
    year = movie.get_year()
    add(years, [year], movie)

    cast = movie.get_actors()
    add(actors, cast, movie)

    dircs = movie.get_directors()
    add(directors, dircs, movie)

    categs = movie.get_genres()
    add(genres, categs, movie)

# test(actors, actors_test)
# test(genres, genres_test)
# test(years, years_test)
# test(directors, directors_test)

actors_objects = []
directors_object = []
genres_object = []
years_object = []


def create(dictionary, lst, obj_type):
    for obj in dictionary:
        object_ = obj_type(obj)
        object_.set_movies(dictionary[obj])
        lst.append(object_)


create(actors, actors_objects, Actor)
create(directors, directors_object, Director)
create(years, years_object, Year)
create(genres, genres_object, Genre)


def calculate(lst):
    type_ = lst[0].type_
    result.write(f' ----------   {type_}   ----------\n')
    lst.sort(key=lambda x: x.get_rate()['me'], reverse=True)
    result.write(f"The Best 10 {type_} For Me Are : \n")
    for i in range(10):
        result.write(f'\t{i+1} - {lst[i].name} With Rate: Me --> {lst[i].get_rate()["me"]} -- IMDb -->{lst[i].get_rate()["IMDb"]} \n\t\tHis Best 3 Movies Are : {lst[i].get_best()}\n\t\t')
        result.write(f'And His Worst 3 Movies Are : {lst[i].get_worst()}\n')
        result.write('\n')
    result.write('\n')
    
    lst.sort(key=lambda x: x.get_rate()['me'], reverse=False)
    result.write(f"The Worst 10 {type_} For Me Are : \n")
    for i in range(10):
        result.write(f'\t{i+1} - {lst[i].name} With Rate: Me --> {lst[i].get_rate()["me"]} -- IMDb -->{lst[i].get_rate()["IMDb"]}\n\t\tHis Best 3 Movies Are : {lst[i].get_best()}\n\t\t')
        result.write(f'And His Worst 3 Movies Are : {lst[i].get_worst()}\n')
        result.write('\n')
    result.write('\n')
        
    lst.sort(key=lambda x: len(x), reverse=True)
    result.write(f"The Most 10 {type_} I Have Seen Movies For Are : \n")
    for i in range(10):
        result.write(f'\t{i+1} - {lst[i].name} With Rate: Me --> {lst[i].get_rate()["me"]} -- IMDb -->{lst[i].get_rate()["IMDb"]} And {len(lst[i])} Movies\n\t\tHis Best 3 Movies Are : {lst[i].get_best()}\n\t\t')
        result.write(f'And His Worst 3 Movies Are : {lst[i].get_worst()}\n')
        result.write('\n')
    result.write('\n\n')

    lst.sort(key=lambda x: (len(x) * x.get_rate()['me']), reverse=True)
    result2.write(f"The Best 10 {type_} For Me Are : \n")
    for i in range(10):
        result2.write(f'\t{i+1} - {lst[i].name} With Rate: Me --> {lst[i].get_rate()["me"]} -- IMDb -->{lst[i].get_rate()["IMDb"]} And Total Rate Of --> {round(len(lst[i])*lst[i].get_rate()["me"],2)}\n\t\tHis Best 3 Movies Are : {lst[i].get_best()}\n\t\t')
        result2.write(f'And His Worst 3 Movies Are : {lst[i].get_worst()}\n')
        result2.write('\n')
    result2.write('\n')
    lst.sort(key=lambda x: (len(x) * x.get_rate()['me']), reverse=False)
    result2.write(f"The Worst 10 {type_} For Me Are : \n")
    for i in range(10):
        result2.write(f'\t{i+1} - {lst[i].name} With Rate: Me --> {lst[i].get_rate()["me"]} -- IMDb -->{lst[i].get_rate()["IMDb"]} And Total Rate Of --> {round(len(lst[i])*lst[i].get_rate()["me"],2)}\n\t\tHis Best 3 Movies Are : {lst[i].get_best()}\n\t\t')
        result2.write(f'And His Worst 3 Movies Are : {lst[i].get_worst()}\n')
        result2.write('\n')
    result2.write('\n')


calculate(actors_objects)
calculate(directors_object)
calculate(years_object)
calculate(genres_object)

result.close()
result2.close()
actors_test.close()
directors_test.close()
genres_test.close()
years_test.close()






