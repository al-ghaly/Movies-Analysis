from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import openpyxl
import time
import os

os.chdir(r"C:\Users\mohamed alghaly\Desktop")

username = "Your email"
password = "Your password"

workbook = openpyxl.load_workbook('Cast.xlsx')
sheet = workbook.create_sheet(title="sheet")

data = {}
Browser = webdriver.Firefox()
Browser.get(r'https://www.imdb.com')
sign_in = Browser.find_element_by_css_selector(r'.imdb-header__signin-text')
sign_in.click()
apple = Browser.find_element_by_css_selector(r'a.list-group-item:nth-child(5)')
apple.click()
name = Browser.find_element_by_id('account_name_text_field')
name.send_keys(username)
name.send_keys(Keys.ENTER)
time.sleep(10)
password = Browser.find_element_by_id("password_text_field")
password.send_keys(password)
password.send_keys(Keys.ENTER)
time.sleep(20)
watch_list = Browser.find_element_by_css_selector(r'.sc-ckVGcZ > a:nth-child(1) > div:nth-child(2)')
watch_list.click()
while True:
    try:
        expand = Browser.find_element_by_css_selector('.load-more')
        expand.click()
    except:
        break
movies = Browser.find_element_by_css_selector('.lister-details')
print(movies.text)
movies = int(movies.text[:3])
time.sleep(10)
for i in range(movies):
    actors = 1
    try:
        name = Browser.find_element_by_css_selector(f'div.series:nth-child({i+1}) > div:nth-child(1) > div:nth-child(2) > h3:nth-child(1) > a:nth-child(1)')
        name = name.text
    except:
        try:
            name = Browser.find_element_by_css_selector(f'div.lister-item:nth-child({i+1}) > div:nth-child(1) > div:nth-child(2) > h3:nth-child(1) > a:nth-child(1)')
            name = name.text
        except:
            name = "UNKNOWN"
    data[name] = []
    print("working on : ", name)
    while True:
        try:
            actor = Browser.find_element_by_css_selector(f'div.lister-item:nth-child({i+1}) > div:nth-child(1) > div:nth-child(2) > div:nth-child(4) > a:nth-child({actors})')
            data[name].append(actor.text)
            print("Actor : ", actor.text)
            actors += 2
        except:
            break

k = 1
for j in data:
    print(f"The movie : {j} With lead actors : {data[j]}")
    sheet.cell(k, 1).value = j
    sheet.cell(k, 2).value = ",".join(data[j])
    k += 1
workbook.save("Cast.xlsx")






